import io
import os
import re
import csv
from collections import OrderedDict
from typing import List, Tuple, Dict

import fitz  # PyMuPDF
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import Response, HTMLResponse, FileResponse

try:
    import openpyxl  # requires openpyxl in requirements.txt
except Exception:
    openpyxl = None


app = FastAPI(title="PDF → CSV (артикул / наименование / всего / категория)", version="3.6.3")

# -------------------------
# Regex
# -------------------------
RX_SIZE = re.compile(r"\b\d{2,}[xх×]\d{2,}(?:[xх×]\d{1,})?\b", re.IGNORECASE)
RX_MM = re.compile(r"мм", re.IGNORECASE)
RX_WEIGHT = re.compile(r"\b\d+(?:[.,]\d+)?\s*кг\.?\b", re.IGNORECASE)

RX_MONEY_LINE = re.compile(r"^\d+(?:[ \u00a0]\d{3})*(?:[.,]\d+)?\s*₽$")
RX_INT = re.compile(r"^\d+$")
RX_ANY_RUB = re.compile(r"₽")

RX_DIMS_ANYWHERE = re.compile(
    r"\s*\d{1,4}[xх×]\d{1,4}(?:[xх×]\d{1,5})?\s*мм\.?\s*",
    re.IGNORECASE,
)


def normalize_space(s: str) -> str:
    s = (s or "").replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalize_key(name: str) -> str:
    s = normalize_space(name).lower()
    s = s.replace("×", "x").replace("х", "x")
    s = RX_DIMS_ANYWHERE.sub(" ", s)
    s = normalize_space(s)
    return s


def strip_dims_anywhere(name: str) -> str:
    name = normalize_space(name)
    name2 = RX_DIMS_ANYWHERE.sub(" ", name)
    return normalize_space(name2)


# -------------------------
# Артикулы (Art.xlsx)
# -------------------------
def load_article_map() -> Tuple[Dict[str, str], str]:
    if openpyxl is None:
        return {}, "openpyxl_not_installed"

    path = os.getenv("ART_XLSX_PATH", "Art.xlsx")
    if not os.path.exists(path):
        return {}, f"file_not_found:{path}"

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        return {}, f"cannot_open:{e}"

    header = [normalize_space(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
    товар_col = 1
    арт_col = 2
    for idx, h in enumerate(header, start=1):
        if h.lower() == "товар":
            товар_col = idx
        if h.lower() == "артикул":
            арт_col = idx

    m: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        товар = ws.cell(r, товар_col).value
        арт = ws.cell(r, арт_col).value
        if not товар or not арт:
            continue
        товар_s = normalize_space(str(товар))
        арт_s = normalize_space(str(арт))
        if not товар_s or not арт_s:
            continue
        m[normalize_key(товар_s)] = арт_s

    return m, "ok"


ARTICLE_MAP, ARTICLE_MAP_STATUS = load_article_map()

CATEGORY_VALUE = 2

# Картинка-инструкция (положи рядом с main.py)
INSTRUCTION_IMAGE_PATH = os.getenv("INSTRUCTION_IMAGE_PATH", "instruction.jpg")


# -------------------------
# PDF parsing
# -------------------------
def split_lines(page: fitz.Page) -> List[str]:
    txt = page.get_text("text") or ""
    lines = [normalize_space(x) for x in txt.splitlines()]
    return [x for x in lines if x]


def is_noise(line: str) -> bool:
    low = (line or "").strip().lower()
    if not low:
        return True
    if low.startswith("страница:"):
        return True
    if low.startswith("ваш проект"):
        return True
    if "проект создан" in low:
        return True
    if "развертка стены" in low:
        return True
    if "стоимость проекта" in low:
        return True
    return False


def is_totals_block(line: str) -> bool:
    low = (line or "").strip().lower()
    return (
        low.startswith("общий вес")
        or low.startswith("максимальный габарит заказа")
        or low.startswith("адрес:")
        or low.startswith("телефон:")
        or low.startswith("email")
    )


def is_project_total_only(line: str) -> bool:
    return bool(re.fullmatch(r"\d+\s*₽", normalize_space(line)))


def is_header_token(line: str) -> bool:
    low = normalize_space(line).lower().replace("–", "-").replace("—", "-")
    return low in {"фото", "товар", "габариты", "вес", "цена за шт", "кол-во", "сумма"}


def looks_like_dim_or_weight(line: str) -> bool:
    if RX_WEIGHT.search(line):
        return True
    if RX_SIZE.search(line) and RX_MM.search(line):
        return True
    return False


def looks_like_money_or_qty(line: str) -> bool:
    if RX_MONEY_LINE.fullmatch(line):
        return True
    if RX_INT.fullmatch(line):
        return True
    return False


def clean_name_from_buffer(buf: List[str]) -> str:
    filtered = []
    for ln in buf:
        if is_noise(ln) or is_header_token(ln) or is_totals_block(ln) or is_project_total_only(ln):
            continue
        filtered.append(ln)

    while filtered and (looks_like_dim_or_weight(filtered[-1]) or looks_like_money_or_qty(filtered[-1])):
        filtered.pop()

    name = normalize_space(" ".join(filtered))
    name = re.sub(r"^Фото\s*", "", name, flags=re.IGNORECASE).strip()
    name = re.sub(r"^Товар\s*", "", name, flags=re.IGNORECASE).strip()
    name = strip_dims_anywhere(name)
    return name


def parse_items(pdf_bytes: bytes) -> Tuple[List[Tuple[str, int]], Dict]:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")

    ordered = OrderedDict()
    buf: List[str] = []
    in_totals = False

    stats = {
        "pages": 0,
        "items_found": 0,
        "article_map_size": len(ARTICLE_MAP),
        "article_map_status": ARTICLE_MAP_STATUS,
    }

    for page in doc:
        stats["pages"] += 1
        lines = split_lines(page)
        if not lines:
            continue

        i = 0
        while i < len(lines):
            line = lines[i]

            if is_noise(line) or is_header_token(line):
                i += 1
                continue

            if is_project_total_only(line) or is_totals_block(line):
                in_totals = True
                buf.clear()
                i += 1
                continue

            if in_totals:
                i += 1
                continue

            # anchor: money -> qty -> money
            if RX_MONEY_LINE.fullmatch(line):
                end = min(len(lines), i + 10)

                qty_idx = None
                for j in range(i + 1, end):
                    if RX_INT.fullmatch(lines[j]):
                        q = int(lines[j])
                        if 1 <= q <= 500:
                            qty_idx = j
                            break

                if qty_idx is None:
                    buf.append(line)
                    i += 1
                    continue

                sum_idx = None
                for j in range(qty_idx + 1, end):
                    if RX_MONEY_LINE.fullmatch(lines[j]) or RX_ANY_RUB.search(lines[j]):
                        sum_idx = j
                        break

                if sum_idx is None:
                    buf.append(line)
                    i += 1
                    continue

                name = clean_name_from_buffer(buf)
                buf.clear()

                if name:
                    qty = int(lines[qty_idx])
                    ordered[name] = ordered.get(name, 0) + qty
                    stats["items_found"] += 1

                i = sum_idx + 1
                continue

            buf.append(line)
            i += 1

    return list(ordered.items()), stats


# -------------------------
# CSV output (Excel-friendly)
# -------------------------
def make_csv_excel_friendly(rows: List[Tuple[str, int]]) -> bytes:
    out = io.StringIO()
    writer = csv.writer(
        out,
        delimiter=";",
        quotechar='"',
        quoting=csv.QUOTE_MINIMAL,
        lineterminator="\r\n",
    )

    writer.writerow(["Артикул", "Наименование", "Всего", "Категория"])

    for name, qty in rows:
        art = ARTICLE_MAP.get(normalize_key(name), "")
        writer.writerow([art, name, qty, CATEGORY_VALUE])

    return out.getvalue().encode("utf-8-sig")  # UTF-8 BOM


# -------------------------
# UI (миниатюра + открыть полностью)
# -------------------------
HOME_HTML = "\n".join([
    "<!doctype html>",
    "<html lang='ru'>",
    "<head>",
    "  <meta charset='utf-8' />",
    "  <meta name='viewport' content='width=device-width, initial-scale=1' />",
    "  <title>PDF → CSV</title>",
    "  <style>",
    "    :root { --bg:#0b0f17; --card:#121a2a; --text:#e9eefc; --muted:#a8b3d6; --border:rgba(255,255,255,.08); --btn:#4f7cff; }",
    "    body { margin:0; font-family: system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;",
    "           background: radial-gradient(1200px 600px at 20% 10%, #18234a 0%, var(--bg) 55%); color: var(--text); }",
    "    .wrap { min-height: 100vh; display:flex; align-items:center; justify-content:center; padding: 28px; }",
    "    .card { width:min(900px, 100%); background: rgba(18,26,42,.92); border: 1px solid var(--border);",
    "            border-radius: 18px; padding: 22px; box-shadow: 0 18px 60px rgba(0,0,0,.45); }",
    "    .top { display:flex; gap:14px; align-items:center; justify-content:space-between; flex-wrap:wrap; }",
    "    h1 { margin:0; font-size: 28px; letter-spacing: .2px; }",
    "    .hint { margin: 8px 0 0; color: var(--muted); font-size: 14px; }",
    "    .badge { font-size: 12px; color: var(--muted); border: 1px solid var(--border); padding: 6px 10px; border-radius: 999px; }",
    "    .row { margin-top: 18px; display:flex; gap: 12px; align-items:center; flex-wrap:wrap; }",
    "    .file { display:flex; align-items:center; gap:10px; padding: 10px 12px; border: 1px dashed var(--border);",
    "            border-radius: 14px; background: rgba(255,255,255,.02); }",
    "    button { padding: 10px 14px; border: 0; border-radius: 14px; cursor: pointer; font-weight: 800;",
    "             background: var(--btn); color: #0b1020; }",
    "    button:disabled { opacity: .55; cursor:not-allowed; }",
    "    .status { margin-top: 14px; font-size: 14px; color: var(--muted); white-space: pre-wrap; }",
    "    .status.ok { color: #79ffa8; }",
    "    .status.err { color: #ff7b8a; }",
    "    .help { margin-top: 16px; }",
    "    .helphead { display:flex; align-items:center; justify-content:space-between; gap:10px; flex-wrap:wrap; }",
    "    .helptitle { font-weight: 700; color: var(--text); }",
    "    .openfull { font-size: 13px; color: var(--muted); text-decoration: underline; cursor: pointer; }",
    "    .thumb { margin-top: 10px; border: 1px solid var(--border); border-radius: 14px; overflow:hidden;",
    "             background: rgba(255,255,255,.02); cursor: zoom-in; }",
    "    .thumb img { display:block; width:100%; height:auto; max-height: 260px; object-fit: cover; object-position: top; }",
    "    /* modal */",
    "    .modal { position: fixed; inset: 0; background: rgba(0,0,0,.75); display:none; align-items:center; justify-content:center; padding: 18px; }",
    "    .modal.open { display:flex; }",
    "    .modalcard { width:min(1200px, 100%); background: rgba(18,26,42,.96); border: 1px solid var(--border);",
    "                 border-radius: 18px; overflow:hidden; box-shadow: 0 30px 100px rgba(0,0,0,.6); }",
    "    .modalbar { display:flex; align-items:center; justify-content:space-between; padding: 10px 12px; border-bottom: 1px solid var(--border); }",
    "    .modalbar .t { color: var(--text); font-weight: 700; font-size: 14px; }",
    "    .close { background: transparent; color: var(--muted); border: 1px solid var(--border);",
    "             border-radius: 12px; padding: 8px 10px; cursor:pointer; font-weight: 700; }",
    "    .modalbody { background: #0b0f17; }",
    "    .modalbody img { display:block; width:100%; height:auto; }",
    "  </style>",
    "</head>",
    "<body>",
    "  <div class='wrap'>",
    "    <div class='card'>",
    "      <div class='top'>",
    "        <div>",
    "          <h1>PDF → CSV</h1>",
    "          <div class='hint'>Загрузите PDF и скачайте CSV для импорта.</div>",
    "        </div>",
    "        <div class='badge'>CSV: ; • UTF-8 • BOM</div>",
    "      </div>",
    "      <div class='row'>",
    "        <div class='file'>",
    "          <input id='pdf' type='file' accept='application/pdf,.pdf' />",
    "        </div>",
    "        <button id='btn' disabled>Скачать CSV</button>",
    "      </div>",
    "      <div id='status' class='status'></div>",
    "",
    "      <div class='help' id='help' style='display:none;'>",
    "        <div class='helphead'>",
    "          <div class='helptitle'>Мини-инструкция</div>",
    "          <div class='openfull' id='openfull'>Открыть полностью</div>",
    "        </div>",
    "        <div class='thumb' id='thumb'>",
    "          <img src='/instruction.jpg' alt='Инструкция' />",
    "        </div>",
    "      </div>",
    "    </div>",
    "  </div>",
    "",
    "  <div class='modal' id='modal' aria-hidden='true'>",
    "    <div class='modalcard'>",
    "      <div class='modalbar'>",
    "        <div class='t'>Инструкция</div>",
    "        <button class='close' id='close'>Закрыть</button>",
    "      </div>",
    "      <div class='modalbody'>",
    "        <img src='/instruction.jpg' alt='Инструкция (полный размер)' />",
    "      </div>",
    "    </div>",
    "  </div>",
    "",
    "  <script>",
    "    const input = document.getElementById('pdf');",
    "    const btn = document.getElementById('btn');",
    "    const statusEl = document.getElementById('status');",
    "    const help = document.getElementById('help');",
    "    const thumb = document.getElementById('thumb');",
    "    const openfull = document.getElementById('openfull');",
    "    const modal = document.getElementById('modal');",
    "    const closeBtn = document.getElementById('close');",
    "",
    "    // покажем блок помощи, если картинка доступна",
    "    fetch('/instruction.jpg', { method: 'HEAD' }).then(r => { if (r.ok) help.style.display = 'block'; });",
    "",
    "    function ok(msg){ statusEl.className='status ok'; statusEl.textContent=msg; }",
    "    function err(msg){ statusEl.className='status err'; statusEl.textContent=msg; }",
    "    function neutral(msg){ statusEl.className='status'; statusEl.textContent=msg||''; }",
    "",
    "    function openModal(){ modal.classList.add('open'); modal.setAttribute('aria-hidden','false'); }",
    "    function closeModal(){ modal.classList.remove('open'); modal.setAttribute('aria-hidden','true'); }",
    "",
    "    thumb.addEventListener('click', openModal);",
    "    openfull.addEventListener('click', openModal);",
    "    closeBtn.addEventListener('click', closeModal);",
    "    modal.addEventListener('click', (e) => { if (e.target === modal) closeModal(); });",
    "    document.addEventListener('keydown', (e) => { if (e.key === 'Escape') closeModal(); });",
    "",
    "    input.addEventListener('change', () => {",
    "      const f = input.files && input.files[0];",
    "      btn.disabled = !f;",
    "      neutral(f ? ('Выбран файл: ' + f.name) : '');",
    "    });",
    "",
    "    btn.addEventListener('click', async () => {",
    "      const f = input.files && input.files[0];",
    "      if (!f) return;",
    "      btn.disabled = true;",
    "      neutral('Обработка…');",
    "      try {",
    "        const fd = new FormData();",
    "        fd.append('file', f);",
    "        const resp = await fetch('/extract', { method: 'POST', body: fd });",
    "        if (!resp.ok) {",
    "          let text = await resp.text();",
    "          try { const j = JSON.parse(text); if (j.detail) text = String(j.detail); } catch(e) {}",
    "          throw new Error(text || ('HTTP ' + resp.status));",
    "        }",
    "        const blob = await resp.blob();",
    "        const base = (f.name || 'items.pdf').replace(/\\.pdf$/i, '');",
    "        const filename = base + '.csv';",
    "        const url = URL.createObjectURL(blob);",
    "        const a = document.createElement('a');",
    "        a.href = url;",
    "        a.download = filename;",
    "        document.body.appendChild(a);",
    "        a.click();",
    "        a.remove();",
    "        URL.revokeObjectURL(url);",
    "        ok('Готово! Файл скачан: ' + filename);",
    "      } catch(e) {",
    "        err('Ошибка: ' + String(e.message || e));",
    "      } finally {",
    "        btn.disabled = !(input.files && input.files[0]);",
    "      }",
    "    });",
    "  </script>",
    "</body>",
    "</html>",
])


@app.get("/health")
def health():
    return {
        "status": "ok",
        "article_map_size": len(ARTICLE_MAP),
        "article_map_status": ARTICLE_MAP_STATUS,
        "category_value": CATEGORY_VALUE,
        "instruction_image_exists": os.path.exists(INSTRUCTION_IMAGE_PATH),
    }


@app.api_route("/", methods=["GET", "HEAD"], response_class=HTMLResponse)
def home():
    return HOME_HTML


@app.get("/instruction.jpg")
def instruction_image():
    if not os.path.exists(INSTRUCTION_IMAGE_PATH):
        raise HTTPException(status_code=404, detail="instruction.jpg not found")
    return FileResponse(INSTRUCTION_IMAGE_PATH, media_type="image/jpeg")


@app.post("/extract")
async def extract(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Загрузите PDF файл (.pdf).")

    pdf_bytes = await file.read()

    try:
        rows, stats = parse_items(pdf_bytes)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Не удалось распарсить PDF: {e}")

    if not rows:
        raise HTTPException(
            status_code=422,
            detail=f"Не удалось найти позиции по шаблону (деньги ₽ → кол-во → деньги ₽). debug={stats}",
        )

    csv_bytes = make_csv_excel_friendly(rows)
    return Response(
        content=csv_bytes,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="items.csv"'},
    )
