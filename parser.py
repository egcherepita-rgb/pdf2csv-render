from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import pdfplumber
from openpyxl import load_workbook


# -----------------------------
# Нормализация / извлечение кода и цвета
# -----------------------------

_COLOR_PAT = re.compile(
    r"\b(бел(ый|ая|ое|ые)|графит(овый|овая|овое|овые)?|сер(ый|ая|ое|ые)|черн(ый|ая|ое|ые))\b",
    re.IGNORECASE,
)

_CODE_PAT = re.compile(
    r"\b([A-ZА-Я]{1,5}[ -]?\d{1,4}(?:[-/]\d{1,4})?)\b"
)


def _norm(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _norm_key(s: str) -> str:
    s = _norm(s).lower()
    s = s.replace("ё", "е")
    # убираем мусорные знаки, но сохраняем дефисы/цифры/буквы
    s = re.sub(r"[^a-zа-я0-9\-\/ ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _extract_color(text: str) -> Optional[str]:
    m = _COLOR_PAT.search(text or "")
    if not m:
        return None
    c = m.group(0).lower().replace("ё", "е")
    # приводим к “белый/графит/серый/черный”
    if c.startswith("бел"):
        return "белый"
    if c.startswith("графит"):
        return "графит"
    if c.startswith("сер"):
        return "серый"
    if c.startswith("черн"):
        return "черный"
    return c


def _extract_code(text: str) -> Optional[str]:
    """
    Вытаскивает “GR-65”, “GS-150”, “GPd 30x10” и т.п.
    Дальше используется как ключ вместе с цветом.
    """
    if not text:
        return None
    # частный случай: GPd 30x10 (есть буквы+цифры+размер)
    m = re.search(r"\b([A-ZА-Я]{1,5}[a-zа-я]{0,3})\s*(\d{1,4}x\d{1,4})\b", text, re.IGNORECASE)
    if m:
        return (m.group(1) + " " + m.group(2)).upper()

    m = _CODE_PAT.search(text.upper())
    if not m:
        return None
    return m.group(1).replace(" ", "-").upper()


# -----------------------------
# Каталог items.xlsx
# -----------------------------

@dataclass(frozen=True)
class CatalogItem:
    code: str
    color: Optional[str]
    name: str
    article: Optional[str]


def _load_items_xlsx(path: str) -> Dict[Tuple[str, Optional[str]], CatalogItem]:
    """
    Ожидается, что в items.xlsx есть хотя бы столбец с названием (и желательно код/артикул).
    Мы строим индекс по (code, color). Если color нет — индекс по (code, None).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]

    def col(*names: str) -> Optional[int]:
        for n in names:
            if n in header:
                return header.index(n)
        return None

    # Популярные варианты заголовков
    col_name = col("наименование", "name", "номенклатура")
    col_code = col("код", "code", "model", "модель", "обозначение", "позиция")
    col_article = col("артикул", "sku", "article")
    col_color = col("цвет", "color")

    if col_name is None:
        # Без имени каталог бесполезен
        return {}

    idx: Dict[Tuple[str, Optional[str]], CatalogItem] = {}

    for r in rows[1:]:
        if not r or all(v is None or str(v).strip() == "" for v in r):
            continue

        name = _norm(str(r[col_name] or ""))
        if not name:
            continue

        # code: либо из отдельной колонки, либо вытащим из имени
        code_val = str(r[col_code]).strip() if col_code is not None and r[col_code] is not None else ""
        code = _extract_code(code_val) or _extract_code(name)
        if not code:
            continue

        color = None
        if col_color is not None and r[col_color] is not None:
            color = _extract_color(str(r[col_color]))
        if color is None:
            color = _extract_color(name)

        article = str(r[col_article]).strip() if col_article is not None and r[col_article] is not None else None
        item = CatalogItem(code=code, color=color, name=name, article=article)

        idx[(code, color)] = item
        # запасной ключ “без цвета”
        if (code, None) not in idx:
            idx[(code, None)] = item

    return idx


# -----------------------------
# Парсинг PDF → строки (имя, кол-во) в порядке PDF
# -----------------------------

@dataclass
class PdfRow:
    raw_name: str
    qty: float


def _try_extract_tables(pdf: pdfplumber.PDF) -> List[PdfRow]:
    out: List[PdfRow] = []

    table_settings = {
        "vertical_strategy": "lines",
        "horizontal_strategy": "lines",
        "snap_tolerance": 3,
        "join_tolerance": 3,
        "edge_min_length": 20,
        "min_words_vertical": 3,
        "min_words_horizontal": 1,
        "intersection_tolerance": 3,
        "text_tolerance": 2,
    }

    for page in pdf.pages:
        tables = page.extract_tables(table_settings=table_settings) or []
        for t in tables:
            # ищем колонки “Наименование” и “Кол-во”
            # но т.к. PDF разные — пробуем эвристики:
            # 1) строка-шапка содержит “наимен” и “кол”
            if not t or len(t) < 2:
                continue

            header = [(_norm(c or "")).lower() for c in t[0]]
            name_i = None
            qty_i = None

            for i, h in enumerate(header):
                if name_i is None and ("наимен" in h or "номенклат" in h or "товар" in h):
                    name_i = i
                if qty_i is None and ("кол" in h or "кол-" in h or "qty" in h or "количество" in h):
                    qty_i = i

            # если не нашли по шапке — часто наименование в 1-й или 2-й колонке, кол-во ближе к концу
            if name_i is None:
                name_i = 1 if len(header) > 1 else 0
            if qty_i is None:
                qty_i = max(0, len(header) - 2)

            for r in t[1:]:
                if not r or len(r) <= max(name_i, qty_i):
                    continue

                raw_name = _norm(r[name_i] or "")
                raw_qty = _norm(str(r[qty_i] or ""))

                if not raw_name:
                    continue

                # количество: 1, 2.5, 3,00
                m = re.search(r"(\d+(?:[.,]\d+)?)", raw_qty)
                if not m:
                    continue
                qty = float(m.group(1).replace(",", "."))

                out.append(PdfRow(raw_name=raw_name, qty=qty))

    # фильтр от мусора: иногда в таблицу попадают итоги/заголовки
    cleaned: List[PdfRow] = []
    for x in out:
        n = x.raw_name.lower()
        if "итого" in n or "всего" in n or "сумм" in n:
            continue
        cleaned.append(x)

    return cleaned


def _fallback_regex_parse(pdf: pdfplumber.PDF) -> List[PdfRow]:
    """
    Если таблицы не извлеклись, парсим “полосами” по тексту.
    Это не идеально, но лучше чем “смещение количества”.
    """
    text = []
    for page in pdf.pages:
        t = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
        text.append(t)
    big = "\n".join(text)
    big = big.replace("\u00a0", " ")
    big = re.sub(r"[ \t]+", " ", big)

    # Ищем блоки вида:
    # <НАИМЕНОВАНИЕ> ... <цена> ... <кол-во> ... <сумма>
    # Кол-во чаще всего отдельным числом ближе к концу строки.
    rows: List[PdfRow] = []

    # Нормальная эвристика: берем строки, где есть “шт” или где в конце “... <qty> ... ₽”
    for line in big.splitlines():
        line = _norm(line)
        if len(line) < 5:
            continue
        if "₽" not in line and "руб" not in line.lower():
            continue

        # пробуем взять количество как “последнее число перед суммой/валютой”
        nums = re.findall(r"\d+(?:[.,]\d+)?", line)
        if len(nums) < 2:
            continue

        # qty обычно маленькое число (1..999) и НЕ похоже на цену (xxxx.xx)
        # возьмём последнее “целое/почти целое” до конца
        qty = None
        for cand in reversed(nums):
            if "." in cand or "," in cand:
                # дробное количество тоже бывает, но редко; допустим
                pass
            try:
                v = float(cand.replace(",", "."))
            except:
                continue
            if 0 < v <= 9999:
                qty = v
                break
        if qty is None:
            continue

        # имя — всё до первого большого денежного числа или до qty
        # упрощенно: отрежем хвост по валютным значениям
        name = re.split(r"\b\d{2,}[.,]?\d*\b", line, maxsplit=1)[0]
        name = _norm(name)
        if len(name) < 3:
            continue

        rows.append(PdfRow(raw_name=name, qty=qty))

    return rows


def parse_pdf_items(pdf_bytes: bytes) -> List[PdfRow]:
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        rows = _try_extract_tables(pdf)
        if rows:
            return rows
        return _fallback_regex_parse(pdf)


# -----------------------------
# Сбор CSV
# -----------------------------

def build_csv_from_pdf(
    pdf_bytes: bytes,
    items_xlsx_path: str,
    delimiter: str = ";",
    use_catalog: bool = True,
) -> str:
    pdf_rows = parse_pdf_items(pdf_bytes)

    catalog: Dict[Tuple[str, Optional[str]], CatalogItem] = {}
    if use_catalog:
        try:
            catalog = _load_items_xlsx(items_xlsx_path)
        except Exception:
            # если каталог сломан — не падаем, просто выводим как в PDF
            catalog = {}

    output = io.StringIO()
    writer = csv.writer(output, delimiter=delimiter, lineterminator="\n")

    # Формат CSV (можете поменять)
    writer.writerow(["Артикул", "Наименование", "Кол-во"])

    for r in pdf_rows:
        raw_name = _norm(r.raw_name)
        qty = r.qty

        # по умолчанию: как в PDF
        final_name = raw_name
        article = ""

        # если хотим подтягивать из каталога — делаем это ТОЛЬКО аккуратно:
        # ключ: code + color (цвет берём из PDF!)
        if catalog:
            code = _extract_code(raw_name)
            color = _extract_color(raw_name)  # ВАЖНО: цвет именно из PDF
            if code:
                hit = catalog.get((code, color)) or catalog.get((code, None))
                if hit:
                    # не перетираем цвет: если в hit.name другой цвет — оставляем PDF-вариант
                    final_name = hit.name
                    if color and _extract_color(final_name) != color:
                        # гарантируем, что “белый” не станет “графит”
                        # добавим цвет из PDF в конец, если его нет
                        if color not in _norm_key(final_name):
                            final_name = f"{final_name} {color}"
                    article = hit.article or ""

        writer.writerow([article, final_name, _format_qty(qty)])

    return output.getvalue()


def _format_qty(qty: float) -> str:
    # чтобы 2.0 стало "2", а 2.5 осталось "2.5"
    if abs(qty - round(qty)) < 1e-9:
        return str(int(round(qty)))
    s = f"{qty:.6f}".rstrip("0").rstrip(".")
    return s
